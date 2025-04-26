# Hayden Coffey
import sys
import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

def get_pdf_table(file):
    MAGIC_PHRASE = "Well"
    all_tables = []
    with pdfplumber.open(file) as pdf:

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue

                df = pd.DataFrame(table)

                # Check if the MAGIC_PHRASE column exists in the header
                header = df.iloc[0]
                if MAGIC_PHRASE in header.values:
                    df.columns = header  # Set first row as header
                    df = df[1:]  # Drop the header row from data
                    all_tables.append(df)

    # Combine all relevant tables
    final_df = pd.concat(all_tables, ignore_index=True)

    return final_df

def process_pdf_table(table_df, input_ref_df, color_ref_df):
    table_df["Content"] = table_df["Well"].apply(lambda well: get_target_value(well, input_ref_df))
    table_df["Target"] = table_df["Well"].apply(lambda well: get_target_value(well, color_ref_df))

    table_df = table_df.drop(['Fluor'], axis=1)

    return table_df

def process_csv_table(table_df, input_ref_df, color_ref_df):
    table_df["Target"] = table_df["Well"].apply(lambda well: get_target_value(well, color_ref_df))

    # Add Content column after Target column
    col_index = table_df.columns.get_loc('Target') + 1
    content_values = table_df["Well"].apply(lambda well: get_target_value(well, input_ref_df))
    table_df.insert(col_index, "Content", content_values)

    # Optional, drop rows with "Undetermined" Cq values.
    #table_df = table_df.loc[table_df['Cq'] != 'Undetermined']

    return table_df

def get_input_ref_df(file):
    wb = load_workbook(file)

    source = wb.active
    ws = wb.copy_worksheet(source)

    val_list = (list(ws.values))

    i = 0
    for l in val_list:
        if all(value is None for value in l):
            break
        i+=1


    content_ref_df = pd.DataFrame(val_list[0:i]).drop(index=0).set_index(0)

    #Strip whitespace
    content_ref_df.index = content_ref_df.index.astype(str).str.strip()
    #content_ref_df.columns = content_ref_df.columns.astype(str).str.strip()

    # Create color reference table.
    color_ref_df = content_ref_df.copy()

    row = 0
    col = 0

    # Create reference table with theme color values
    # Have to combine tint (fp rounded) with theme as some colors have same theme value.
    for r in ws[2:i]:
        col = 0
        for c in r[1:]:
            if (color_ref_df.iloc[row,col]):
                color_ref_df.iloc[row,col] = c.fill.fgColor.theme +  round(c.fill.fgColor.tint,2)
            col += 1
        row += 1

    # Color row has translations from theme color to string
    color_row = (ws[i+2])

    # Extract key:values of color translation into a table
    color_table = []
    for c in color_row:
        if c.value == None:
            continue

        color_table.append([c.fill.fgColor.theme + round(c.fill.fgColor.tint,2), c.value])

    # assert uniqueness
    themes = [pair[0] for pair in color_table]
    targets = [pair[1] for pair in color_table]
    assert len(themes)  == len(set(themes)),  f"Duplicate theme-elements found: \n{themes}\n{targets}"
    assert len(targets) == len(set(targets)), f"Duplicate target-elements found: \n{themes}\n{targets}"

    # Create color dict and apply to color ref df to create final lookup df.
    color_df = (pd.DataFrame(color_table))
    color_df=color_df.set_index(0)

    color_match_dict = (color_df[1].to_dict())
    color_ref_df = color_ref_df.replace(color_match_dict)

    return content_ref_df, color_ref_df

def get_target_value(well, df_lookup):
    """Extract the lookup value from df_lookup based on a well identifier like 'A1'."""
    # Parse the well string: 
    # Row letter is the first character; column number is the rest (it could be multiple digits)
    row_label = well[0].upper()
    col_number = int(well[1:])

    # Get the corresponding lookup value from df_lookup
    if row_label in df_lookup.index and col_number in df_lookup.columns:
        return df_lookup.loc[row_label, col_number]
    else:
        return ""

def main():

    if len(sys.argv) < 3:
        print("Usage: python {} input.pdf input_ref.xlsx (optional:) output.xlsx".format(sys.argv[0]))
        return

    input_file = sys.argv[1]
    input_ref_file = sys.argv[2]

    output_file = "converted_output.xlsx"

    if len(sys.argv) == 4:
        output_file = sys.argv[3] 

    input_ref_df, color_ref_df = get_input_ref_df(input_ref_file)

    # Process data from either pdf or csv file. 
    if input_file.endswith('.pdf'):
        table_df = get_pdf_table(input_file)
        table_df = process_pdf_table(table_df, input_ref_df, color_ref_df)
    elif input_file.endswith('.csv'):
        table_df = pd.read_csv(input_file)
        table_df = process_csv_table(table_df, input_ref_df, color_ref_df)
    else:
        print("Error: Unsupported input file type. Supported: .pdf, .csv")
        exit()

    if os.path.exists(output_file):
        raise FileExistsError(f"The file '{output_file}' already exists. Will not overwrite.")

    table_df.to_excel(output_file, index=False)

if __name__ == "__main__":
    main()
